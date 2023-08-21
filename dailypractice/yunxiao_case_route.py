import openpyxl


def import_testcase(excel_file, menu='3.15.2.8执行库|功能测试|'):
    # 使用前确保导出的excel用例文件将所有字段全部导出了
    excel_file = excel_file
    work_book = openpyxl.load_workbook(excel_file)
    work_sheet = work_book['testcase items']
    row_num = work_sheet.max_row
    print(row_num)

    for i in range(2, row_num+1):
        # 读取每行的C列“目录”的值后进行改变拓展目录
        content = work_sheet.cell(row=i, column=3).value
        content = menu + content
        # 将修改后的目录重新写入excel文件
        work_sheet.cell(row=i, column=3, value=content)
        # print(content)
    work_book.save(excel_file)

excel_file = r'C:\Users\fu\Downloads\GSLB基线用例-用例优化版本-导出用例-20230625162147.xlsx'
import_testcase(excel_file)