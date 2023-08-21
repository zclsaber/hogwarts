import linecache
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference, Series
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm

filename = r"C:\Users\fu\Downloads\mem_usage.log"
count = 1

wb = Workbook()
ws = wb.active

# 逐行遍历文件中所有文本
while True:
    line = linecache.getline(filename, count)
    if not line:
        break
    if count % 2 == 0:  # 判断奇偶性-偶数
        try:
            ws.append([(line.split(":")[0]+":"+line.split(":")[1]), float(line.split(",")[1].split(" ")[0]), float(line.split(",")[1].split(" ")[1])])
        except Exception:
            pass
    count += 1

wb.save(r"C:\Users\fu\Downloads\mem_usage.xlsx")

# # 重新加载文件获取数据
# wb = load_workbook(r"C:\Users\fu\Downloads\mem_usage.xlsx")
# ws = wb["Sheet"]

# 画曲线图
# font_path = r"C:\Windows\Fonts\msyh.ttc"
# font_prop = fm.FontProperties(fname=font_path)
# plt.rcParams["font.family"] = font_prop.get_name()
#
#
# df = pd.read_excel(r"C:\Users\fu\Downloads\mem_usage.xlsx", usecols=[0, 1], names=["日期", "CPU"])
# x = df["日期"]
# y = df["CPU"]
#
# fig2, ax = plt.subplots(figsize=(12, 9), dpi=300)
# ax.plot(x, y)
#
# ax.xaxis.set_major_locator(plt.MaxNLocator(13))
# plt.xticks(rotation=45)
#
# plt.gcf().autofmt_xdate()
#
# plt.xlabel("日期", fontproperties=font_prop)
# plt.ylabel("CPU")
# plt.title("长稳CPU曲线", fontproperties=font_prop)
#
# plt.show()